import os
import openpyxl

def rename_files(filename,ori_row,ori_col,new_row,new_col,path_to_folder):
    initial_path = Path(path_to_folder)
    path_file_list = os.listdir(path_to_folder)
    wb = openpyxl.load_workbook(filename)
    sheet_interest = wb.sheetnames[0]  # select first sheet
    sheet = wb[sheet_interest]
    max_row = sheet.max_row
    original_filename = []
    new_filenames = []
    for i in range(0, max_row - ori_row + 1):
        original_filename = original_filename + [sheet.cell(row=ori_row + i, column=ori_col).value]
    for j in range(0,max_row - new_row + 1):
        temp_new_name = sheet.cell(row=new_row + j, column=new_col).value
        while temp_new_name in new_filenames:
            temp_new_name = new_path(temp_new_name)
            if temp_new_name in new_filenames:
                continue
            else:
                print('new filename ','"',sheet.cell(row=new_row + j, column=new_col).value,'" ','has been changed to ','"',temp_new_name,'"',sep='')
                break
        new_filenames = new_filenames + [temp_new_name]
    for k in range(0,max_row - ori_row + 1):
        older_name = original_filename[k]
        newer_name = new_filenames[k]
        specific_path = initial_path/older_name
        if older_name in path_file_list:
            specific_path.rename(Path(specific_path.parent, newer_name))
        else:
            print(older_name,'is not in',path_to_folder)
